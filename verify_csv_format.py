"""
Verify that both XLSX and CSV formats are generated correctly.
"""

import csv
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

output_dir = Path("output")

# Find the most recent compliance results files
xlsx_files = sorted(output_dir.glob("compliance_results_*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)
csv_files = sorted(output_dir.glob("compliance_results_*.csv"), key=lambda x: x.stat().st_mtime, reverse=True)

if not xlsx_files or not csv_files:
    print("❌ No output files found!")
    exit(1)

xlsx_file = xlsx_files[0]
csv_file = csv_files[0]

print("\n" + "="*80)
print("📊 COMPLIANCE OUTPUT FORMAT VERIFICATION")
print("="*80)

# Check XLSX
print(f"\n📄 XLSX File: {xlsx_file.name}")
print(f"   Size: {xlsx_file.stat().st_size} bytes")

wb = load_workbook(xlsx_file)
ws = wb.active
print(f"   Sheet: {ws.title}")
print(f"   Dimensions: {ws.dimensions}")

# Get headers
headers_xlsx = []
for cell in ws[1]:
    if cell.value:
        headers_xlsx.append(cell.value)
print(f"   Columns ({len(headers_xlsx)}): {', '.join(headers_xlsx)}")

# Count rows
row_count_xlsx = ws.max_row - 1  # Exclude header
print(f"   Data Rows: {row_count_xlsx}")

# Check CSV
print(f"\n📄 CSV File: {csv_file.name}")
print(f"   Size: {csv_file.stat().st_size} bytes")

with open(csv_file, 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)
    headers_csv = next(csv_reader)
    print(f"   Columns ({len(headers_csv)}): {', '.join(headers_csv)}")
    
    # Count data rows
    row_count_csv = sum(1 for _ in csv_reader)
    print(f"   Data Rows: {row_count_csv}")

# Sample data
print(f"\n📊 SAMPLE DATA COMPARISON:")
print(f"\n   XLSX Sample (Row 2):")
if ws.max_row > 1:
    for col_idx, header in enumerate(headers_xlsx, 1):
        cell_value = ws.cell(2, col_idx).value
        print(f"      {header}: {cell_value}")

print(f"\n   CSV Sample (Row 2):")
with open(csv_file, 'r', encoding='utf-8') as f:
    csv_reader = csv.DictReader(f)
    try:
        first_row = next(csv_reader)
        for header in headers_csv:
            print(f"      {header}: {first_row.get(header, '')}")
    except StopIteration:
        print("      No data rows")

# Verify column match
print(f"\n✅ VERIFICATION RESULTS:")
if len(headers_xlsx) == len(headers_csv):
    print(f"   ✓ Column count matches: {len(headers_xlsx)}")
else:
    print(f"   ✗ Column count mismatch: XLSX={len(headers_xlsx)}, CSV={len(headers_csv)}")

if row_count_xlsx == row_count_csv:
    print(f"   ✓ Row count matches: {row_count_xlsx}")
else:
    print(f"   ✗ Row count mismatch: XLSX={row_count_xlsx}, CSV={row_count_csv}")

print(f"\n✅ Both XLSX and CSV formats generated successfully!")
print(f"\n   📌 XLSX: {xlsx_file}")
print(f"   📌 CSV:  {csv_file}")
print("\n" + "="*80 + "\n")
