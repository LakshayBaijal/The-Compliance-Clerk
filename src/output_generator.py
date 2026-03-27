"""
Output Generator - Creates standalone output.xlsx file with all pipeline results.

This module generates a comprehensive output.xlsx file that consolidates:
- All extracted data from pipeline execution
- Summary statistics
- Performance metrics
- Document type breakdown
- Processing status overview
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

logger = logging.getLogger(__name__)


class OutputGenerator:
    """Generate comprehensive output.xlsx file."""

    # Styling
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        """Initialize output generator."""
        self.workbook = None
        logger.info("Output Generator initialized")

    def generate(
        self,
        batch_results: Dict[str, Any],
        output_path: str = "output.xlsx"
    ) -> str:
        """
        Generate comprehensive output.xlsx file.

        Args:
            batch_results: Dictionary from process_batch() containing:
                - results: List of extraction result rows
                - summary: Summary statistics
                - output_excel: Path to Excel file
            output_path: Path where to save output.xlsx (default: output.xlsx)

        Returns:
            Path to generated output.xlsx file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        results = batch_results.get("results", [])
        summary = batch_results.get("summary", {})

        # Create sheets in order
        self._create_overview_sheet(results, summary)
        self._create_detailed_results_sheet(results)
        self._create_echallan_sheet(results)
        self._create_na_permission_sheet(results)
        self._create_statistics_sheet(results, summary)
        self._create_processing_log_sheet(results)

        # Save file
        self.workbook.save(output_path)
        logger.info(f"Output.xlsx generated at: {output_path}")

        return str(output_path)

    def _create_overview_sheet(self, results: List[Dict], summary: Dict):
        """Create executive overview sheet."""
        sheet = self.workbook.create_sheet("Overview", 0)

        # Title
        sheet["A1"] = "COMPLIANCE DOCUMENT EXTRACTION - OVERVIEW"
        sheet["A1"].font = self.TITLE_FONT
        sheet.merge_cells("A1:F1")

        # Timestamp
        sheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet.merge_cells("A2:F2")

        # Key Metrics
        row = 4
        sheet[f"A{row}"] = "KEY METRICS"
        sheet[f"A{row}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row}:F{row}")

        row += 2
        metrics = [
            ("Total Pages Processed", summary.get("total_pages", 0)),
            ("Successful Extractions", summary.get("success", 0)),
            ("Partial Extractions", summary.get("partial", 0)),
            ("Failed Extractions", summary.get("failed", 0)),
            ("Success Rate", f"{summary.get('success_rate', 0)}%"),
            ("Total Tokens Used", summary.get("total_tokens", 0)),
            ("Processing Time (s)", summary.get("elapsed_seconds", 0)),
        ]

        sheet[f"A{row}"] = "Metric"
        sheet[f"B{row}"] = "Value"
        sheet[f"A{row}"].font = self.HEADER_FONT
        sheet[f"B{row}"].font = self.HEADER_FONT
        sheet[f"A{row}"].fill = self.HEADER_FILL
        sheet[f"B{row}"].fill = self.HEADER_FILL

        row += 1
        for metric_name, metric_value in metrics:
            sheet[f"A{row}"] = metric_name
            sheet[f"B{row}"] = metric_value
            sheet[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Document Type Breakdown
        row += 2
        sheet[f"A{row}"] = "DOCUMENT TYPE BREAKDOWN"
        sheet[f"A{row}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row}:C{row}")

        row += 1
        doc_types = self._count_document_types(results)

        sheet[f"A{row}"] = "Type"
        sheet[f"B{row}"] = "Count"
        sheet[f"C{row}"] = "Percentage"
        for col in ["A", "B", "C"]:
            sheet[f"{col}{row}"].font = self.HEADER_FONT
            sheet[f"{col}{row}"].fill = self.HEADER_FILL

        row += 1
        total_docs = sum(doc_types.values())
        for doc_type, count in sorted(doc_types.items(), key=lambda x: x[1], reverse=True):
            sheet[f"A{row}"] = doc_type
            sheet[f"B{row}"] = count
            sheet[f"C{row}"] = f"{(count/total_docs*100):.1f}%" if total_docs > 0 else "0%"
            sheet[f"B{row}"].alignment = Alignment(horizontal="right")
            sheet[f"C{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Status Breakdown
        row += 2
        sheet[f"A{row}"] = "STATUS BREAKDOWN"
        sheet[f"A{row}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row}:C{row}")

        row += 1
        status_counts = self._count_status(results)

        sheet[f"A{row}"] = "Status"
        sheet[f"B{row}"] = "Count"
        sheet[f"C{row}"] = "Percentage"
        for col in ["A", "B", "C"]:
            sheet[f"{col}{row}"].font = self.HEADER_FONT
            sheet[f"{col}{row}"].fill = self.HEADER_FILL

        row += 1
        total_status = sum(status_counts.values())
        for status, count in sorted(status_counts.items(), key=lambda x: x[1], reverse=True):
            sheet[f"A{row}"] = status.upper()
            sheet[f"B{row}"] = count
            sheet[f"C{row}"] = f"{(count/total_status*100):.1f}%" if total_status > 0 else "0%"

            # Color code status
            if status == "success":
                sheet[f"A{row}"].fill = self.SUCCESS_FILL
            elif status == "partial":
                sheet[f"A{row}"].fill = self.WARNING_FILL
            elif status == "failed":
                sheet[f"A{row}"].fill = self.ERROR_FILL

            sheet[f"B{row}"].alignment = Alignment(horizontal="right")
            sheet[f"C{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Adjust column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15

    def _create_detailed_results_sheet(self, results: List[Dict]):
        """Create detailed results sheet with all extraction data."""
        sheet = self.workbook.create_sheet("Detailed Results", 1)

        # Headers
        headers = [
            "File Name",
            "Page",
            "Type",
            "Status",
            "Confidence",
            "Method",
            "Issues",
            "Tokens",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, result in enumerate(results, 2):
            sheet.cell(row=row_idx, column=1).value = result.get("file_name", "")
            sheet.cell(row=row_idx, column=2).value = result.get("page_number", "")
            sheet.cell(row=row_idx, column=3).value = result.get("document_type", "")
            sheet.cell(row=row_idx, column=4).value = result.get("status", "").upper()
            sheet.cell(row=row_idx, column=5).value = result.get("confidence", 0)
            sheet.cell(row=row_idx, column=6).value = result.get("extraction_method", "")
            sheet.cell(row=row_idx, column=7).value = "; ".join(result.get("validation_issues", []))
            sheet.cell(row=row_idx, column=8).value = result.get("tokens_used", 0)

            # Color code status
            status = result.get("status", "")
            if status == "success":
                sheet.cell(row=row_idx, column=4).fill = self.SUCCESS_FILL
            elif status == "partial":
                sheet.cell(row=row_idx, column=4).fill = self.WARNING_FILL
            elif status == "failed":
                sheet.cell(row=row_idx, column=4).fill = self.ERROR_FILL

            # Alignment
            sheet.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
            sheet.cell(row=row_idx, column=8).alignment = Alignment(horizontal="right")

        # Adjust column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 8
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 12
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 15
        sheet.column_dimensions["G"].width = 30
        sheet.column_dimensions["H"].width = 10

    def _create_echallan_sheet(self, results: List[Dict]):
        """Create eChallan-specific results sheet."""
        sheet = self.workbook.create_sheet("eChallan Results", 2)

        # Filter eChallan records
        echallan_records = [r for r in results if r.get("document_type") == "ECHALLAN"]

        if not echallan_records:
            sheet["A1"] = "No eChallan records found"
            return

        # Headers
        headers = [
            "File Name",
            "Page",
            "Vehicle Type",
            "License Plate",
            "Violation",
            "Fine Amount",
            "Date",
            "Status",
            "Confidence",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, record in enumerate(echallan_records, 2):
            data = record.get("echallan_data", {})
            sheet.cell(row=row_idx, column=1).value = record.get("file_name", "")
            sheet.cell(row=row_idx, column=2).value = record.get("page_number", "")
            sheet.cell(row=row_idx, column=3).value = data.get("vehicle_type", "")
            sheet.cell(row=row_idx, column=4).value = data.get("license_plate", "")
            sheet.cell(row=row_idx, column=5).value = data.get("violation_type", "")
            sheet.cell(row=row_idx, column=6).value = data.get("fine_amount", "")
            sheet.cell(row=row_idx, column=7).value = data.get("violation_date", "")
            sheet.cell(row=row_idx, column=8).value = record.get("status", "").upper()
            sheet.cell(row=row_idx, column=9).value = record.get("confidence", 0)

            # Color code status
            status = record.get("status", "")
            if status == "success":
                sheet.cell(row=row_idx, column=8).fill = self.SUCCESS_FILL
            elif status == "partial":
                sheet.cell(row=row_idx, column=8).fill = self.WARNING_FILL

        # Adjust column widths
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 15

    def _create_na_permission_sheet(self, results: List[Dict]):
        """Create NA Permission-specific results sheet."""
        sheet = self.workbook.create_sheet("NA Permission Results", 3)

        # Filter NA Permission records
        na_records = [r for r in results if r.get("document_type") == "NA_PERMISSION"]

        if not na_records:
            sheet["A1"] = "No NA Permission records found"
            return

        # Headers
        headers = [
            "File Name",
            "Page",
            "Property ID",
            "Lessee Name",
            "Lessor Name",
            "Property Area",
            "Agreement Date",
            "Lease Period",
            "Status",
            "Confidence",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, record in enumerate(na_records, 2):
            data = record.get("na_data", {})
            sheet.cell(row=row_idx, column=1).value = record.get("file_name", "")
            sheet.cell(row=row_idx, column=2).value = record.get("page_number", "")
            sheet.cell(row=row_idx, column=3).value = data.get("property_id", "")
            sheet.cell(row=row_idx, column=4).value = data.get("lessee_name", "")
            sheet.cell(row=row_idx, column=5).value = data.get("lessor_name", "")
            sheet.cell(row=row_idx, column=6).value = data.get("property_area", "")
            sheet.cell(row=row_idx, column=7).value = data.get("agreement_date", "")
            sheet.cell(row=row_idx, column=8).value = data.get("lease_period", "")
            sheet.cell(row=row_idx, column=9).value = record.get("status", "").upper()
            sheet.cell(row=row_idx, column=10).value = record.get("confidence", 0)

            # Color code status
            status = record.get("status", "")
            if status == "success":
                sheet.cell(row=row_idx, column=9).fill = self.SUCCESS_FILL
            elif status == "partial":
                sheet.cell(row=row_idx, column=9).fill = self.WARNING_FILL

        # Adjust column widths
        for col in range(1, 11):
            sheet.column_dimensions[get_column_letter(col)].width = 15

    def _create_statistics_sheet(self, results: List[Dict], summary: Dict):
        """Create statistical analysis sheet."""
        sheet = self.workbook.create_sheet("Statistics", 4)

        row = 1
        sheet[f"A{row}"] = "EXTRACTION STATISTICS"
        sheet[f"A{row}"].font = self.TITLE_FONT
        sheet.merge_cells(f"A{row}:B{row}")

        row += 2
        sheet[f"A{row}"] = "Metric"
        sheet[f"B{row}"] = "Value"
        sheet[f"A{row}"].font = self.HEADER_FONT
        sheet[f"B{row}"].font = self.HEADER_FONT
        sheet[f"A{row}"].fill = self.HEADER_FILL
        sheet[f"B{row}"].fill = self.HEADER_FILL

        row += 1
        stats_data = [
            ("Total Pages", summary.get("total_pages", 0)),
            ("Success Count", summary.get("success", 0)),
            ("Partial Count", summary.get("partial", 0)),
            ("Failed Count", summary.get("failed", 0)),
            ("Success Rate (%)", summary.get("success_rate", 0)),
            ("Average Confidence", self._calculate_avg_confidence(results)),
            ("Total Tokens", summary.get("total_tokens", 0)),
            ("Processing Time (s)", summary.get("elapsed_seconds", 0)),
            ("Pages per Second", self._calculate_speed(summary)),
        ]

        for metric, value in stats_data:
            sheet[f"A{row}"] = metric
            sheet[f"B{row}"] = value
            sheet[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        row += 2
        sheet[f"A{row}"] = "EXTRACTION METHOD BREAKDOWN"
        sheet[f"A{row}"].font = self.TITLE_FONT
        sheet.merge_cells(f"A{row}:B{row}")

        row += 1
        sheet[f"A{row}"] = "Method"
        sheet[f"B{row}"] = "Count"
        sheet[f"A{row}"].font = self.HEADER_FONT
        sheet[f"B{row}"].font = self.HEADER_FONT
        sheet[f"A{row}"].fill = self.HEADER_FILL
        sheet[f"B{row}"].fill = self.HEADER_FILL

        row += 1
        methods = self._count_extraction_methods(results)
        for method, count in sorted(methods.items(), key=lambda x: x[1], reverse=True):
            sheet[f"A{row}"] = method
            sheet[f"B{row}"] = count
            sheet[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15

    def _create_processing_log_sheet(self, results: List[Dict]):
        """Create processing log sheet."""
        sheet = self.workbook.create_sheet("Processing Log", 5)

        headers = [
            "File Name",
            "Page",
            "Type",
            "Status",
            "Method",
            "Issues Count",
            "Confidence",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_idx, record in enumerate(results, 2):
            sheet.cell(row=row_idx, column=1).value = record.get("file_name", "")
            sheet.cell(row=row_idx, column=2).value = record.get("page_number", "")
            sheet.cell(row=row_idx, column=3).value = record.get("document_type", "")
            sheet.cell(row=row_idx, column=4).value = record.get("status", "").upper()
            sheet.cell(row=row_idx, column=5).value = record.get("extraction_method", "")
            sheet.cell(row=row_idx, column=6).value = len(record.get("validation_issues", []))
            sheet.cell(row=row_idx, column=7).value = record.get("confidence", 0)

            # Color code
            status = record.get("status", "")
            if status == "success":
                sheet.cell(row=row_idx, column=4).fill = self.SUCCESS_FILL
            elif status == "partial":
                sheet.cell(row=row_idx, column=4).fill = self.WARNING_FILL
            elif status == "failed":
                sheet.cell(row=row_idx, column=4).fill = self.ERROR_FILL

        for col in range(1, 8):
            sheet.column_dimensions[get_column_letter(col)].width = 15

    def _count_document_types(self, results: List[Dict]) -> Dict[str, int]:
        """Count documents by type."""
        counts = {}
        for result in results:
            doc_type = result.get("document_type", "UNKNOWN")
            counts[doc_type] = counts.get(doc_type, 0) + 1
        return counts

    def _count_status(self, results: List[Dict]) -> Dict[str, int]:
        """Count results by status."""
        counts = {}
        for result in results:
            status = result.get("status", "unknown")
            counts[status] = counts.get(status, 0) + 1
        return counts

    def _count_extraction_methods(self, results: List[Dict]) -> Dict[str, int]:
        """Count by extraction method."""
        counts = {}
        for result in results:
            method = result.get("extraction_method", "unknown")
            counts[method] = counts.get(method, 0) + 1
        return counts

    def _calculate_avg_confidence(self, results: List[Dict]) -> float:
        """Calculate average confidence."""
        if not results:
            return 0.0
        total = sum(r.get("confidence", 0) for r in results)
        return round(total / len(results), 3)

    def _calculate_speed(self, summary: Dict) -> float:
        """Calculate pages per second."""
        elapsed = summary.get("elapsed_seconds", 1)
        total = summary.get("total_pages", 0)
        if elapsed > 0:
            return round(total / elapsed, 2)
        return 0.0
