"""
Specialized Excel exporter for expected compliance document extraction format.

Generates a single sheet with extracted data in the format:
Sr.no | Village | Survey No. | Area in NA Order | Dated | NA Order No. | 
Lease Deed Doc. No. | Lease Area | Lease Start | ...
"""

import logging
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ComplianceExporter:
    """Exports extraction results in the expected compliance document format."""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Expected columns for Lease Deed (NA Permission)
    NA_PERMISSION_COLUMNS = [
        "Sr.no.",
        "Village",
        "Survey No.",
        "Area in NA Order",
        "Dated",
        "NA Order No.",
        "Lease Deed Doc. No.",
        "Lease Area",
        "Lease Start",
        "Lease End",
        "Status",
    ]

    # Expected columns for Challan
    ECHALLAN_COLUMNS = [
        "Sr.no.",
        "Vehicle Reg.",
        "Challan No.",
        "Violation",
        "Amount Due",
        "Date Issued",
        "Due Date",
        "Status",
        "Payment Status",
    ]

    def __init__(self):
        """Initialize compliance exporter."""
        logger.info("Compliance Exporter initialized")

    def export_compliance_format(
        self,
        results: List[Dict[str, Any]],
        output_path: Path,
    ) -> Path:
        """
        Export extraction results in expected compliance format.

        Args:
            results: List of extraction result dictionaries
            output_path: Path to write Excel file

        Returns:
            Path to created Excel file
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Extracted Data"

        # Separate results by document type
        na_permission_results = [r for r in results if r.get("document_type") == "NA_PERMISSION"]
        echallan_results = [r for r in results if r.get("document_type") == "ECHALLAN"]

        row = 1

        # Export NA Permission data first
        if na_permission_results:
            row = self._export_na_permission_data(worksheet, na_permission_results, row)

        # Add spacing if both types present
        if na_permission_results and echallan_results:
            row += 2

        # Export eChallan data
        if echallan_results:
            row = self._export_echallan_data(worksheet, echallan_results, row)

        # Auto-fit columns
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        workbook.save(output_path)
        logger.info(f"Exported compliance format to {output_path}")

        return output_path

    def _export_na_permission_data(
        self, worksheet, results: List[Dict[str, Any]], start_row: int
    ) -> int:
        """Export NA Permission (Lease Deed) data."""
        row = start_row

        # Write headers
        for col_num, column_title in enumerate(self.NA_PERMISSION_COLUMNS, 1):
            cell = worksheet.cell(row=row, column=col_num)
            cell.value = column_title
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.BORDER

        row += 1

        # Write data rows
        sr_no = 1
        for result in results:
            na_data = result.get("na_data", {})
            validated_data = result.get("validated_data", na_data)

            # Extract values from result
            values = [
                sr_no,  # Sr.no.
                self._extract_village(result.get("file_name", "")),  # Village
                self._extract_survey_no(validated_data),  # Survey No.
                validated_data.get("property_area"),  # Area in NA Order
                validated_data.get("permission_date"),  # Dated
                validated_data.get("property_id"),  # NA Order No.
                validated_data.get("lease_deed_number"),  # Lease Deed Doc. No.
                validated_data.get("property_area"),  # Lease Area
                validated_data.get("permission_date"),  # Lease Start
                validated_data.get("expiry_date"),  # Lease End
                result.get("status", "").upper(),  # Status
            ]

            # Write row
            for col_num, value in enumerate(values, 1):
                cell = worksheet.cell(row=row, column=col_num)
                cell.value = value
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Format status column with colors
                if col_num == len(values):  # Status column
                    if value == "SUCCESS":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif value == "PARTIAL":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    elif value == "FAILED":
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )

            row += 1
            sr_no += 1

        return row

    def _export_echallan_data(
        self, worksheet, results: List[Dict[str, Any]], start_row: int
    ) -> int:
        """Export eChallan data."""
        row = start_row

        # Write headers
        for col_num, column_title in enumerate(self.ECHALLAN_COLUMNS, 1):
            cell = worksheet.cell(row=row, column=col_num)
            cell.value = column_title
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.BORDER

        row += 1

        # Write data rows
        sr_no = 1
        for result in results:
            echallan_data = result.get("echallan_data", {})
            validated_data = result.get("validated_data", echallan_data)

            # Extract values from result
            values = [
                sr_no,  # Sr.no.
                validated_data.get("vehicle_reg_number"),  # Vehicle Reg.
                validated_data.get("challan_number"),  # Challan No.
                validated_data.get("violation_description"),  # Violation
                validated_data.get("amount_due"),  # Amount Due
                validated_data.get("issuing_date"),  # Date Issued
                validated_data.get("payment_due_date"),  # Due Date
                validated_data.get("payment_status", "").upper(),  # Status
                result.get("status", "").upper(),  # Payment Status
            ]

            # Write row
            for col_num, value in enumerate(values, 1):
                cell = worksheet.cell(row=row, column=col_num)
                cell.value = value
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Format status columns with colors
                if col_num in [8, 9]:  # Status columns
                    if value == "SUCCESS":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif value == "PARTIAL":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    elif value == "FAILED":
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )

            row += 1
            sr_no += 1

        return row

    @staticmethod
    def _extract_village(filename: str) -> str:
        """Extract village name from filename."""
        if "rampura" in filename.lower():
            return "Rampura Mota"
        elif "mota" in filename.lower():
            return "Rampura Mota"
        return "Unknown"

    @staticmethod
    def _extract_survey_no(data: Dict) -> str:
        """Extract survey number from extracted data."""
        plot_num = data.get("plot_number", "")
        if plot_num:
            # Extract just the number from "S.No.-256" format
            return plot_num.replace("S.No.-", "").replace("S.No.", "")
        return ""
