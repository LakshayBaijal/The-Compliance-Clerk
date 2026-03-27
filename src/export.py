"""
Excel export module for batch extraction results.

Generates multi-sheet workbooks with extraction results, validation reports, and summaries.
"""

import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports extraction results to formatted Excel workbooks."""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        """Initialize Excel exporter."""
        self.workbook = None
        logger.info("Excel Exporter initialized")

    def export_batch_results(
        self,
        results: List[Dict[str, Any]],
        output_path: Path,
        include_validation: bool = True,
        include_tokens: bool = True,
    ) -> Path:
        """
        Export batch extraction results to Excel workbook.

        Args:
            results: List of extraction result dictionaries
            output_path: Path to write Excel file
            include_validation: Include validation issues sheet
            include_tokens: Include token usage sheet

        Returns:
            Path to created Excel file
        """
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(results)
        self._create_echallan_sheet(results)
        self._create_na_permission_sheet(results)

        if include_validation:
            self._create_validation_sheet(results)

        if include_tokens:
            self._create_token_sheet(results)

        # Save workbook
        self.workbook.save(output_path)
        logger.info(f"Exported batch results to {output_path}")

        return output_path

    def _create_summary_sheet(self, results: List[Dict[str, Any]]):
        """Create summary statistics sheet."""
        sheet = self.workbook.create_sheet("Summary", 0)

        # Title
        sheet["A1"] = "Extraction Batch Summary"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet.merge_cells("A1:D1")

        # Export timestamp
        sheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet.merge_cells("A2:D2")

        # Statistics
        row = 4
        stats = self._calculate_stats(results)

        # Headers
        sheet[f"A{row}"] = "Metric"
        sheet[f"B{row}"] = "Value"
        sheet[f"A{row}"].font = self.HEADER_FONT
        sheet[f"B{row}"].font = self.HEADER_FONT
        sheet[f"A{row}"].fill = self.HEADER_FILL
        sheet[f"B{row}"].fill = self.HEADER_FILL

        row += 1

        # Data rows
        for metric, value in stats.items():
            sheet[f"A{row}"] = metric
            sheet[f"B{row}"] = value
            row += 1

        # Column widths
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 20

    def _create_echallan_sheet(self, results: List[Dict[str, Any]]):
        """Create eChallan extraction results sheet."""
        sheet = self.workbook.create_sheet("eChallan Results")

        # Filter eChallan results
        echallan_results = [
            r for r in results if r.get("document_type") == "ECHALLAN"
        ]

        if not echallan_results:
            sheet["A1"] = "No eChallan documents found"
            return

        # Headers
        headers = [
            "File Name",
            "Page",
            "Challan #",
            "Vehicle Reg",
            "Violation",
            "Amount Due",
            "Payment Status",
            "Due Date",
            "Confidence",
            "Status",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        row = 2
        for result in echallan_results:
            data = result.get("echallan_data", {})
            validated = result.get("validated_data", {})

            sheet.cell(row=row, column=1).value = result.get("file_name", "")
            sheet.cell(row=row, column=2).value = result.get("page_number", "")
            sheet.cell(row=row, column=3).value = (
                validated.get("challan_number") or data.get("challan_number", "")
            )
            sheet.cell(row=row, column=4).value = (
                validated.get("vehicle_reg_number")
                or data.get("vehicle_reg_number", "")
            )
            sheet.cell(row=row, column=5).value = (
                validated.get("violation_description")
                or data.get("violation_description", "")
            )
            sheet.cell(row=row, column=6).value = (
                validated.get("amount_due") or data.get("amount_due", "")
            )
            sheet.cell(row=row, column=7).value = (
                validated.get("payment_status") or data.get("payment_status", "")
            )
            sheet.cell(row=row, column=8).value = (
                validated.get("payment_due_date") or data.get("payment_due_date", "")
            )
            sheet.cell(row=row, column=9).value = round(result.get("confidence", 0), 3)

            # Status color coding
            status = result.get("status", "")
            status_cell = sheet.cell(row=row, column=10)
            status_cell.value = status

            if status == "success":
                status_cell.fill = self.SUCCESS_FILL
            elif status == "partial":
                status_cell.fill = self.WARNING_FILL
            else:
                status_cell.fill = self.ERROR_FILL

            row += 1

        # Adjust column widths
        for i, header in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(i)].width = 15

    def _create_na_permission_sheet(self, results: List[Dict[str, Any]]):
        """Create NA Permission extraction results sheet."""
        sheet = self.workbook.create_sheet("NA Permission Results")

        # Filter NA Permission results
        na_results = [
            r for r in results if r.get("document_type") == "NA_PERMISSION"
        ]

        if not na_results:
            sheet["A1"] = "No NA Permission documents found"
            return

        # Headers
        headers = [
            "File Name",
            "Page",
            "Property ID",
            "Owner Name",
            "Area",
            "Permission Date",
            "Expiry Date",
            "Authority",
            "Status",
            "Confidence",
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        row = 2
        for result in na_results:
            data = result.get("na_data", {})
            validated = result.get("validated_data", {})

            sheet.cell(row=row, column=1).value = result.get("file_name", "")
            sheet.cell(row=row, column=2).value = result.get("page_number", "")
            sheet.cell(row=row, column=3).value = (
                validated.get("property_id") or data.get("property_id", "")
            )
            sheet.cell(row=row, column=4).value = (
                validated.get("owner_name") or data.get("owner_name", "")
            )
            sheet.cell(row=row, column=5).value = (
                validated.get("property_area") or data.get("property_area", "")
            )
            sheet.cell(row=row, column=6).value = (
                validated.get("permission_date") or data.get("permission_date", "")
            )
            sheet.cell(row=row, column=7).value = (
                validated.get("expiry_date") or data.get("expiry_date", "")
            )
            sheet.cell(row=row, column=8).value = (
                validated.get("issuing_authority")
                or data.get("issuing_authority", "")
            )

            # Status color coding
            status = result.get("status", "")
            status_cell = sheet.cell(row=row, column=9)
            status_cell.value = status

            if status == "success":
                status_cell.fill = self.SUCCESS_FILL
            elif status == "partial":
                status_cell.fill = self.WARNING_FILL
            else:
                status_cell.fill = self.ERROR_FILL

            sheet.cell(row=row, column=10).value = round(result.get("confidence", 0), 3)

            row += 1

        # Adjust column widths
        for i, header in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(i)].width = 15

    def _create_validation_sheet(self, results: List[Dict[str, Any]]):
        """Create validation issues sheet."""
        sheet = self.workbook.create_sheet("Validation Issues")

        # Filter results with validation issues
        issues_results = [
            r
            for r in results
            if r.get("validation_issues") and len(r.get("validation_issues", [])) > 0
        ]

        if not issues_results:
            sheet["A1"] = "No validation issues found"
            return

        # Headers
        headers = ["File Name", "Page", "Document Type", "Issue", "Severity"]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        # Data rows
        row = 2
        for result in issues_results:
            for issue in result.get("validation_issues", []):
                sheet.cell(row=row, column=1).value = result.get("file_name", "")
                sheet.cell(row=row, column=2).value = result.get("page_number", "")
                sheet.cell(row=row, column=3).value = result.get("document_type", "")
                sheet.cell(row=row, column=4).value = issue
                sheet.cell(row=row, column=5).value = "Warning"
                sheet.cell(row=row, column=5).fill = self.WARNING_FILL
                row += 1

        # Adjust column widths
        for i in range(1, 6):
            sheet.column_dimensions[get_column_letter(i)].width = 20

    def _create_token_sheet(self, results: List[Dict[str, Any]]):
        """Create token usage sheet."""
        sheet = self.workbook.create_sheet("Token Usage")

        # Collect token data
        token_data = []
        total_tokens = 0

        for result in results:
            tokens = result.get("tokens_used", 0)
            if tokens > 0:
                token_data.append(
                    {
                        "file": result.get("file_name", ""),
                        "page": result.get("page_number", ""),
                        "method": result.get("extraction_method", ""),
                        "tokens": tokens,
                        "confidence": round(result.get("confidence", 0), 3),
                    }
                )
                total_tokens += tokens

        if not token_data:
            sheet["A1"] = "No token usage data"
            return

        # Summary
        sheet["A1"] = "Token Usage Summary"
        sheet["A1"].font = Font(bold=True, size=12)
        sheet["A2"] = f"Total Tokens: {total_tokens}"
        sheet["A3"] = f"Total Extractions: {len(token_data)}"
        sheet["A4"] = f"Average Tokens per Extraction: {round(total_tokens / len(token_data), 1)}"

        # Headers
        row = 6
        headers = ["File Name", "Page", "Method", "Tokens", "Confidence"]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL

        # Data rows
        row = 7
        for data in token_data:
            sheet.cell(row=row, column=1).value = data["file"]
            sheet.cell(row=row, column=2).value = data["page"]
            sheet.cell(row=row, column=3).value = data["method"]
            sheet.cell(row=row, column=4).value = data["tokens"]
            sheet.cell(row=row, column=5).value = data["confidence"]
            row += 1

        # Adjust column widths
        for i in range(1, 6):
            sheet.column_dimensions[get_column_letter(i)].width = 15

    @staticmethod
    def _calculate_stats(results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate batch statistics."""
        total = len(results)
        successful = len([r for r in results if r.get("status") == "success"])
        partial = len([r for r in results if r.get("status") == "partial"])
        failed = len([r for r in results if r.get("status") == "failed"])

        echallan_count = len([r for r in results if r.get("document_type") == "ECHALLAN"])
        na_count = len([r for r in results if r.get("document_type") == "NA_PERMISSION"])

        total_tokens = sum(r.get("tokens_used", 0) for r in results)
        avg_confidence = (
            sum(r.get("confidence", 0) for r in results) / total if total > 0 else 0
        )

        total_issues = sum(len(r.get("validation_issues", [])) for r in results)

        return {
            "Total Extractions": total,
            "Successful": successful,
            "Partial": partial,
            "Failed": failed,
            "eChallan Documents": echallan_count,
            "NA Permission Documents": na_count,
            "Average Confidence": round(avg_confidence, 3),
            "Total Tokens Used": total_tokens,
            "Total Validation Issues": total_issues,
        }
