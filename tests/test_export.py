"""
Tests for Excel Export module.
"""

import sys
import tempfile
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.export import ExcelExporter


def test_excel_exporter_init():
    """Test ExcelExporter initialization."""
    print("Running Excel Export tests...\n")

    exporter = ExcelExporter()
    assert exporter is not None
    print("[PASS] ExcelExporter initialized")


def test_export_empty_results():
    """Test exporting empty results."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        result = exporter.export_batch_results([], output_path)

        assert result.exists()
        print(f"[PASS] Empty batch exported to {result.name}")


def test_export_with_echallan_results():
    """Test exporting eChallan results."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test1.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "extraction_method": "deterministic",
                "confidence": 0.86,
                "status": "success",
                "tokens_used": 0,
                "echallan_data": {
                    "challan_number": "CH001",
                    "vehicle_reg_number": "MH-02-AB-1234",
                    "violation_description": "Speeding",
                    "amount_due": 2000.0,
                    "payment_status": "pending",
                    "payment_due_date": "2026-04-27",
                },
                "validated_data": {
                    "challan_number": "CH001",
                    "vehicle_reg_number": "MH-02-AB-1234",
                    "amount_due": 2000.0,
                },
                "validation_issues": [],
            }
        ]

        result = exporter.export_batch_results(results, output_path)

        assert result.exists()
        # Verify workbook has sheets
        from openpyxl import load_workbook
        wb = load_workbook(result)
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "eChallan Results" in sheet_names
        print(f"[PASS] eChallan results exported with {len(sheet_names)} sheets")


def test_export_with_na_permission_results():
    """Test exporting NA Permission results."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test2.pdf",
                "page_number": 1,
                "document_type": "NA_PERMISSION",
                "extraction_method": "deterministic",
                "confidence": 0.87,
                "status": "success",
                "tokens_used": 0,
                "na_data": {
                    "property_id": "PROP001",
                    "owner_name": "John Doe",
                    "property_area": 5000.0,
                    "permission_date": "2024-01-01",
                    "expiry_date": "2026-01-01",
                    "issuing_authority": "Municipal Authority",
                },
                "validated_data": {
                    "property_id": "PROP001",
                    "property_area": 5000.0,
                },
                "validation_issues": [],
            }
        ]

        result = exporter.export_batch_results(results, output_path)

        assert result.exists()
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert "NA Permission Results" in wb.sheetnames
        print(f"[PASS] NA Permission results exported")


def test_export_with_mixed_results():
    """Test exporting mixed document types."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test1.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "extraction_method": "deterministic",
                "confidence": 0.86,
                "status": "success",
                "tokens_used": 0,
                "echallan_data": {"challan_number": "CH001", "amount_due": 2000.0},
                "validated_data": {"challan_number": "CH001"},
                "validation_issues": [],
            },
            {
                "file_name": "test2.pdf",
                "page_number": 1,
                "document_type": "NA_PERMISSION",
                "extraction_method": "llm",
                "confidence": 0.75,
                "status": "partial",
                "tokens_used": 150,
                "na_data": {"property_id": "PROP001", "property_area": 5000.0},
                "validated_data": {"property_id": "PROP001"},
                "validation_issues": ["Invalid date format"],
            },
        ]

        result = exporter.export_batch_results(results, output_path)

        assert result.exists()
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert len(wb.sheetnames) >= 4  # Summary, eChallan, NA, Validation
        print(f"[PASS] Mixed batch exported with {len(wb.sheetnames)} sheets")


def test_export_with_validation_issues():
    """Test exporting results with validation issues."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test_bad.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "extraction_method": "deterministic",
                "confidence": 0.50,
                "status": "partial",
                "tokens_used": 0,
                "echallan_data": {"amount_due": "invalid"},
                "validated_data": {},
                "validation_issues": [
                    "Invalid amount format",
                    "Missing vehicle registration",
                ],
            }
        ]

        result = exporter.export_batch_results(
            results, output_path, include_validation=True
        )

        assert result.exists()
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert "Validation Issues" in wb.sheetnames

        # Check Validation Issues sheet has data
        val_sheet = wb["Validation Issues"]
        assert val_sheet["A2"].value is not None  # Should have data rows
        print("[PASS] Validation issues sheet populated")


def test_export_with_token_usage():
    """Test exporting token usage data."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test1.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "extraction_method": "llm",
                "confidence": 0.75,
                "status": "success",
                "tokens_used": 125,
                "echallan_data": {},
                "validated_data": {},
                "validation_issues": [],
            },
            {
                "file_name": "test2.pdf",
                "page_number": 1,
                "document_type": "NA_PERMISSION",
                "extraction_method": "llm",
                "confidence": 0.80,
                "status": "success",
                "tokens_used": 170,
                "na_data": {},
                "validated_data": {},
                "validation_issues": [],
            },
        ]

        result = exporter.export_batch_results(
            results, output_path, include_tokens=True
        )

        assert result.exists()
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert "Token Usage" in wb.sheetnames

        # Check Token Usage sheet
        token_sheet = wb["Token Usage"]
        assert "Total Tokens: 295" in str(token_sheet["A2"].value)
        print("[PASS] Token usage sheet with aggregates")


def test_export_summary_sheet():
    """Test summary sheet generation."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": f"test{i}.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN" if i % 2 == 0 else "NA_PERMISSION",
                "extraction_method": "deterministic",
                "confidence": 0.75 + (i * 0.01),
                "status": "success" if i % 2 == 0 else "partial",
                "tokens_used": i * 50,
                "echallan_data": {},
                "na_data": {},
                "validated_data": {},
                "validation_issues": [],
            }
            for i in range(3)
        ]

        result = exporter.export_batch_results(results, output_path)

        from openpyxl import load_workbook
        wb = load_workbook(result)
        summary_sheet = wb["Summary"]

        # Verify summary content
        assert summary_sheet["A1"].value == "Extraction Batch Summary"
        assert "Total Extractions" in str(summary_sheet["A5"].value)
        print("[PASS] Summary sheet generated with statistics")


def test_calculate_stats():
    """Test statistics calculation."""
    exporter = ExcelExporter()

    results = [
        {
            "document_type": "ECHALLAN",
            "status": "success",
            "confidence": 0.86,
            "tokens_used": 0,
            "validation_issues": [],
        },
        {
            "document_type": "NA_PERMISSION",
            "status": "partial",
            "confidence": 0.75,
            "tokens_used": 100,
            "validation_issues": ["Issue 1"],
        },
        {
            "document_type": "ECHALLAN",
            "status": "success",
            "confidence": 0.90,
            "tokens_used": 0,
            "validation_issues": [],
        },
    ]

    stats = exporter._calculate_stats(results)

    assert stats["Total Extractions"] == 3
    assert stats["Successful"] == 2
    assert stats["Partial"] == 1
    assert stats["eChallan Documents"] == 2
    assert stats["NA Permission Documents"] == 1
    assert stats["Total Tokens Used"] == 100
    assert stats["Total Validation Issues"] == 1
    print(f"[PASS] Statistics calculated: {stats['Total Extractions']} extractions, "
          f"{stats['Total Tokens Used']} tokens")


def test_export_without_optional_sheets():
    """Test export with optional sheets disabled."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        exporter = ExcelExporter()

        results = [
            {
                "file_name": "test.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "extraction_method": "deterministic",
                "confidence": 0.86,
                "status": "success",
                "tokens_used": 0,
                "echallan_data": {},
                "validated_data": {},
                "validation_issues": [],
            }
        ]

        result = exporter.export_batch_results(
            results, output_path, include_validation=False, include_tokens=False
        )

        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert "Validation Issues" not in wb.sheetnames
        assert "Token Usage" not in wb.sheetnames
        assert "Summary" in wb.sheetnames
        print("[PASS] Export with optional sheets disabled")


if __name__ == "__main__":
    test_excel_exporter_init()
    test_export_empty_results()
    test_export_with_echallan_results()
    test_export_with_na_permission_results()
    test_export_with_mixed_results()
    test_export_with_validation_issues()
    test_export_with_token_usage()
    test_export_summary_sheet()
    test_calculate_stats()
    test_export_without_optional_sheets()

    print("\n[SUCCESS] All Excel export tests passed!")
