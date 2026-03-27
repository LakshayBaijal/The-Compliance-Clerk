"""
Tests for output_generator module.
"""

import pytest
from pathlib import Path
from src.output_generator import OutputGenerator


@pytest.fixture
def sample_batch_results():
    """Generate sample batch results."""
    return {
        "results": [
            {
                "file_name": "test_echallan.pdf",
                "page_number": 1,
                "document_type": "ECHALLAN",
                "status": "success",
                "confidence": 0.95,
                "extraction_method": "deterministic",
                "validation_issues": [],
                "tokens_used": 0,
                "echallan_data": {
                    "vehicle_type": "Two Wheeler",
                    "license_plate": "GJ-01-AB-1234",
                    "violation_type": "Speeding",
                    "fine_amount": "500",
                    "violation_date": "2026-03-20",
                },
                "na_data": {},
            },
            {
                "file_name": "test_na.pdf",
                "page_number": 1,
                "document_type": "NA_PERMISSION",
                "status": "partial",
                "confidence": 0.65,
                "extraction_method": "ocr",
                "validation_issues": ["Missing lessor name"],
                "tokens_used": 150,
                "na_data": {
                    "property_id": "GJ-001-2026",
                    "lessee_name": "John Doe",
                    "lessor_name": "",
                    "property_area": "1000 sqft",
                    "agreement_date": "2026-03-15",
                    "lease_period": "12 months",
                },
                "echallan_data": {},
            },
            {
                "file_name": "test_unknown.pdf",
                "page_number": 1,
                "document_type": "UNKNOWN",
                "status": "failed",
                "confidence": 0.0,
                "extraction_method": "none",
                "validation_issues": ["Could not determine document type"],
                "tokens_used": 0,
                "na_data": {},
                "echallan_data": {},
            },
        ],
        "summary": {
            "total_pages": 3,
            "success": 1,
            "partial": 1,
            "failed": 1,
            "success_rate": 33.33,
            "total_tokens": 150,
            "elapsed_seconds": 2.5,
        },
    }


def test_output_generator_init():
    """Test OutputGenerator initialization."""
    generator = OutputGenerator()
    assert generator is not None
    assert generator.workbook is None


def test_output_generator_generate(tmp_path, sample_batch_results):
    """Test output generation."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_output.xlsx"

    result = generator.generate(sample_batch_results, str(output_path))

    assert Path(result).exists()
    assert Path(result).suffix == ".xlsx"


def test_output_generator_sheets(tmp_path, sample_batch_results):
    """Test that all required sheets are created."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_sheets.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    # Load workbook to check sheets
    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    expected_sheets = [
        "Overview",
        "Detailed Results",
        "eChallan Results",
        "NA Permission Results",
        "Statistics",
        "Processing Log",
    ]

    for sheet_name in expected_sheets:
        assert sheet_name in wb.sheetnames


def test_output_generator_overview_data(tmp_path, sample_batch_results):
    """Test overview sheet contains correct data."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_overview.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet = wb["Overview"]

    # Check for key metrics
    assert "KEY METRICS" in [cell.value for row in sheet.iter_rows() for cell in row]
    assert "DOCUMENT TYPE BREAKDOWN" in [cell.value for row in sheet.iter_rows() for cell in row]
    assert "STATUS BREAKDOWN" in [cell.value for row in sheet.iter_rows() for cell in row]


def test_output_generator_detailed_results(tmp_path, sample_batch_results):
    """Test detailed results sheet."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_detailed.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet = wb["Detailed Results"]

    # Check headers
    headers = [cell.value for cell in sheet[1]]
    assert "File Name" in headers
    assert "Page" in headers
    assert "Status" in headers


def test_output_generator_echallan_sheet(tmp_path, sample_batch_results):
    """Test eChallan results sheet."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_echallan.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet = wb["eChallan Results"]

    # Check headers
    headers = [cell.value for cell in sheet[1]]
    assert "License Plate" in headers
    assert "Violation" in headers
    assert "Fine Amount" in headers


def test_output_generator_na_permission_sheet(tmp_path, sample_batch_results):
    """Test NA Permission results sheet."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_na.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet = wb["NA Permission Results"]

    # Check headers
    headers = [cell.value for cell in sheet[1]]
    assert "Property ID" in headers
    assert "Lessee Name" in headers
    assert "Lessor Name" in headers


def test_output_generator_statistics(tmp_path, sample_batch_results):
    """Test statistics sheet."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_stats.xlsx"

    generator.generate(sample_batch_results, str(output_path))

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    sheet = wb["Statistics"]

    # Check for statistics section
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "EXTRACTION STATISTICS" in values
    assert "Total Pages" in values


def test_output_generator_empty_results(tmp_path):
    """Test with empty results."""
    generator = OutputGenerator()
    output_path = tmp_path / "test_empty.xlsx"

    batch_results = {
        "results": [],
        "summary": {
            "total_pages": 0,
            "success": 0,
            "partial": 0,
            "failed": 0,
            "success_rate": 0.0,
            "total_tokens": 0,
            "elapsed_seconds": 0,
        },
    }

    result = generator.generate(batch_results, str(output_path))
    assert Path(result).exists()


def test_output_generator_default_path(tmp_path, sample_batch_results, monkeypatch):
    """Test with default output path."""
    monkeypatch.chdir(tmp_path)

    generator = OutputGenerator()
    result = generator.generate(sample_batch_results)

    assert Path(result).exists()
    assert result == "output.xlsx"
