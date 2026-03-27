"""
Verify the compliance output format
"""
from openpyxl import load_workbook
from pathlib import Path

# Find the most recent compliance output file
output_files = list(Path("output").glob("compliance_results_*.xlsx"))
if not output_files:
    print("No output files found!")
    exit(1)

latest_file = max(output_files, key=lambda x: x.stat().st_mtime)
print(f"Checking: {latest_file}\n")

wb = load_workbook(latest_file)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}\n")

print("=" * 120)
print("HEADERS:")
print("=" * 120)
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    headers.append(cell_value)
    print(f"  Column {col}: {cell_value}")

print("\n" + "=" * 120)
print("DATA SAMPLE (First 5 rows):")
print("=" * 120)
for row in range(2, min(7, ws.max_row + 1)):
    row_data = []
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        row_data.append(value)
    print(f"\nRow {row}:")
    for col_num, (header, value) in enumerate(zip(headers, row_data), 1):
        print(f"  {header}: {value}")

print("\n" + "=" * 120)
print(f"Total rows (including header): {ws.max_row}")
print(f"Total columns: {ws.max_column}")
print("=" * 120)
