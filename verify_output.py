#!/usr/bin/env python
from openpyxl import load_workbook

wb = load_workbook("output/output.xlsx")
print("Sheets in output.xlsx:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = ws.max_row
    cols = ws.max_column
    print(f"  - {sheet}: {rows} rows x {cols} columns")
